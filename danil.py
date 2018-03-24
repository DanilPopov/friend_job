from openpyxl import load_workbook
from find_job_posts import find_job_posts

def read_excel_posts(filename):
    work_book = load_workbook(filename)
    work_sheet = work_book['Лист1']

    excel_data = list()
    for row in range(1, work_sheet.max_row+1):
        excel_row = dict()
        if work_sheet.cell(row=row, column=1).value is None:
            break
        excel_row['id'] = work_sheet.cell(row=row, column=1).value
        excel_row['post'] = work_sheet.cell(row=row, column=2).value
        excel_data.append(excel_row)

    return excel_data

test_posts = read_excel_posts('Тестовые посты.xlsx')
keywords_file_1 = 'Ключевые слова.xlsx'

find_job_posts(test_posts, keywords_file_1)


    # post = row['post']
    # post_list.append(post)

# for post in post_list:
    # 1. сделать поиск ключевых слов, чтобы определить, относится ли пост к объявлению о работе. Вернуть True или False
    # 2. сделать поиск ключевых слов, чтобы определить, в каком городе вакансия
    # - искать город в объявлении. 
    # - если нет - искать город у разместившего в профиле
    # 3. сделать поиск ключевых слов, чтобы определить зону ответственности
    # - приводить все ключевые слова к нижнему регистру


