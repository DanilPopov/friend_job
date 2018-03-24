from openpyxl import load_workbook

def find_job_posts(posts, keywords_file):

    def read_excel_is_job_keywords(keywords_file, c):
        work_book = load_workbook(keywords_file)
        work_sheet = work_book['Лист1']

        excel_data = list()
        for row in range(3, work_sheet.max_row+1):
            
            excel_row = dict()
            if work_sheet.cell(row=row, column=c).value is None:
                break
            excel_row['keyword'] = work_sheet.cell(row=row, column=c).value
            excel_data.append(excel_row)

        return excel_data

    is_job_hashtags = read_excel_is_job_keywords(keywords_file, 1) 
    is_job_keys_1 = read_excel_is_job_keywords(keywords_file, 2) 
    is_job_keys_2 = read_excel_is_job_keywords(keywords_file, 3)
    is_job_minus = read_excel_is_job_keywords(keywords_file, 4)


    post_list = []
    is_job_hashtags_list = []
    is_job_keywords_list_1 = []
    is_job_keywords_list_2 = []
    is_job_minus_list = []

    for row in is_job_hashtags:
        is_job_hashtags_list.append(row['keyword'])
    # print(is_job_hashtags_list)

    for row in is_job_keys_1:
        is_job_keywords_list_1.append(row['keyword'])
    # print(is_job_keywords_list_1)

    for row in is_job_keys_2:
        is_job_keywords_list_2.append(row['keyword'])
    # print(is_job_keywords_list_2)

    for row in is_job_minus:
        is_job_minus_list.append(row['keyword'])
    # print(is_job_minus_list)


    def find_keywords_in_post(post_text, keywords_list):

        for key in keywords_list:
            if key in post_text:
                return True
        return False

    def define_if_job(post_text):

        if find_keywords_in_post(post_text, is_job_hashtags_list):
            return True
        else:
            is_in_list_1 = find_keywords_in_post(post_text, is_job_keywords_list_1)
            is_in_list_2 = find_keywords_in_post(post_text, is_job_keywords_list_2)
            not_in_minus_list = not find_keywords_in_post(post_text, is_job_minus_list)
            if all([is_in_list_1, is_in_list_2, not_in_minus_list]):
                return True
        return False

    for row in posts:
        if define_if_job(row['post'].lower()):
            row['is_job'] = True
            # row['city'] = define_city(кщц) # присваивает город к посту с вакансией
            post_list.append(row)


    print(post_list)


    for row in post_list:
        print(row['id'])