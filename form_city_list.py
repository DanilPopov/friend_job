from openpyxl import load_workbook

def read_excel_cities(filename):
    work_book = load_workbook(filename)
    work_sheet = work_book['Лист1']

    excel_data = list()
    for row in range(1, work_sheet.max_row+1):
        case = list()
        excel_row = dict()
        if work_sheet.cell(row=row, column=1).value is None:
            break
        excel_row['city'] = work_sheet.cell(row=row, column=1).value
        excel_row['case1'] = work_sheet.cell(row=row, column=2).value
        excel_row['case2'] = work_sheet.cell(row=row, column=3).value
        excel_row['case3'] = work_sheet.cell(row=row, column=4).value
        excel_row['case4'] = work_sheet.cell(row=row, column=5).value
        excel_row['case5'] = work_sheet.cell(row=row, column=6).value
        excel_row['case6'] = work_sheet.cell(row=row, column=7).value
        case = [excel_row['city'], excel_row['case1'], excel_row['case1'], excel_row['case1'], excel_row['case1'], excel_row['case1'], excel_row['case1']]
        excel_row = {'city': excel_row['city'], 'case': case}
        excel_data.append(excel_row)

    return excel_data

print(read_excel_cities('city_list2.xlsx'))


# with open('city.csv', 'r', encoding='cp1251') as f:
#     fields = ['city', 'population']
#     reader = csv.DictReader(f, fields, delimiter=';')
    
#     city_list = []
#     for row in reader:
#         city_list.append(row['city'])
#     print(city_list)

